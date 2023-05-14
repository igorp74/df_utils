"""
## DATAFRAME UTILS
*#A collection of useful functions for optimized and faster work with DataFrame*

---
👔 by Igor Perkovic

🛠 CREATED: 2020-10-13 08:39:29
📆 CHANGED: 2023-05-14 23:35:52

---
⚙ PREREQUISITES:
📘 Libraries: xlsxwriter, pandas, sqlalchemy

"""
from datetime import datetime, timedelta
from pathlib import Path
import numpy as np
import pandas as pd


# Databases
def df_2_mssqlsrv(df, engine_name, schema_name, table_name, ifexist):
    """
    ====================================

    🏷 Creates or append table in MS SQL Server database
        from DataFrame data

    📌 ARGUMENTS:
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    - df           (DataFrame)
    - engine_name  (SQLAlchemy engine)
    - schema_name  (SQL Server schema name)
    - table_name   (str) Table name
    - ifexist      (str) fail    = just throw an error and stop
                         replace = replace existing table with a new data
                         append  = apeend to existing table


    🎯 RETURNS
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    → SQL Server table data
    """
    print('Inserting into database...')

    # Optimal chunk-size for SQL Server import
    chunk_size=999//(df.shape[1]+1)

    try:
        df.to_sql(name=table_name, con=engine_name, schema=schema_name, index=False, if_exists=ifexist, chunksize=chunk_size, method='multi')
        print(f'Successfully INSERTED table: {table_name}\n')
    except:
        print(f'INSERT FAILED for {table_name}')

def df_2_sqlite(df, db_path, table_name):
    """
    ====================================

    🏷 Creates a SQLite database file from DataFrame

    📌 ARGUMENTS:
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    - df         (DataFrame)
    - db_path    (Path) Database name with path
    - table_name (str)  Table name

    🎯 RETURNS
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    → SQLite file (database)
    """
    from sqlalchemy import create_engine

    engine = create_engine(f'sqlite:///{db_path}', echo=False)
    sqlite_connection = engine.connect()

    try:
        df.to_sql(table_name, sqlite_connection, if_exists='fail')
        print(f'Successfully created database {db_path} and table {table_name}')
    except:
        print('ERROR creating SQLite database !')



# Excel
def get_xlsx_data(fn, sn=''):
    """
    ==============================================

    🏷 Simple wrapper over pd.read_excel function

    📌 ARGUMENTS:
    ――――――――――――――――――――――――――――――――――――――――――――――
    - fn (Path) file  name
    - sb (Str)  sheet name (optional) # if left, first sheet would be used

    🎯 RETURNS
    ――――――――――――――――――――――――――――――――――――――――――――――
    → DataFrame
    """

    try:
        print('Getting the data from xlsx to DataFrame')
        if len(sn):
            df = pd.read_excel(fn, sheet_name = sn)
        else:
            df = pd.read_excel(fn)
        print('Successfully imported!\n')
        return df
    except:
        print('Oh, no! Data Import ERROR!')


def df_merged_headers(cl, delimiter):
    merged_headers = []
    for c in cl:
        tmp = []
        for i in c:
            if ('Unnamed' in str(i)) or ('*' in str(i)) or (i is np.nan) or ('None' in str(i)):
                pass
            else:
                tmp.append(str(i).replace('\n','').strip().upper())
        merged_headers.append(delimiter.join(tmp[0: len(tmp)]))
    return merged_headers


def get_xlsx(fn, **kwargs):
    """
    ==================================================
    🏷 Gets data from one file and multiple worksheets


    📌ARGUMENTS
    ――――――――――――――――――――――――――――――――――――――――――――――――――
    fn (Path)            Existing xlsx file

    sn_list     (list) List of desired existing sheet names
    sn_col      (str)  Column name for column with worksheet names
    header_rows (int)  Number of rows for handling multiple headers
    delimiter   (str)  Delimiter for merged header names
    clean       (Bool) Remove empty columns and rows from dataframes
    collect     (Bool) Collect dataframes of all sheets into list of dataframes
    concatenate (Bool) Concatenate all work sheets
    to_pickle   (Bool) Save each worksheet in a pickle file
    destination (Path) Path for saving pickle files

    🎯 RETURNS
    ――――――――――――――――――――――――――――――――――――――――――――――――――
    → List of DataFrames - one DataFrame from each worksheet
    """

    xlsx = pd.ExcelFile(fn)
    sheets = xlsx.sheet_names
    print(f'Reading data from: {fn}\nFound work sheets: {sheets}\n')

    # Default values
    #-----------------
    sn = []     # sheet names - which I want to read
    work_sheets = []
    snc = ''
    nr = 0
    delimiter = '-'
    src_headers = 0
    collect = 0
    p = 0 # pickle
    c = 0 # clean
    con = 0 # concatenate
    dest_pick = Path('')

    # Get dynamic argument
    for k,v in kwargs.items():
        # Check all arguments
        # ---------------------------
        # print("%s = %s" % (k, v))

        if k == 'collect':
            collect = v
        if k == 'to_pickle':
            p = v
        if k == 'sn_list':
            sn  = v
        if k == 'sn_col':
            snc  = v
        if k == 'header_rows':
            nr = v
        if k == 'delimiter':
            delimiter = v
        if k == 'clean':
            c = v
        if k == 'concatenate':
            con = v
        if k == 'destination':
            dest_pick = v

    if nr > 1:
        src_headers = [i for i in range(nr)]

    dfs = []

    if len(sn):
        work_sheets = sn
    else:
        work_sheets = sheets

    if collect:
        print('Collecting DataFrames to a list is ON\n')

    for ws in work_sheets:
        print('Reading worksheet:',ws)
        if src_headers:
            print('Headers:', src_headers)

        df = xlsx.parse(ws, header=src_headers)

        if len(snc):
            df[snc] = ws

        if nr > 1:
            cl = df.columns.tolist()
            df.columns = df_merged_headers(cl, delimiter)

        # Clean
        if c:
            df = df.dropna(axis=1, how='all')
            df = df.dropna(axis=0, how='all')
            df = df.reset_index(drop=True)

        # Collect DataFrames into dataframe list
        if collect:
            dfs.append(df)

        # Save worksheets to pickle files
        if p > 0 and con == 0:
            tmp_fn = fn.stem + '_' + ws + '.pickle'
            tmp_pn = dest_pick / tmp_fn
            df.to_pickle(tmp_pn)
            print(f'Created: {tmp_pn}\n')

    if collect:
        if con:
            print('\nConcatenating collected dataframes...')
            res = pd.concat(dfs)
            res.columns = res.columns.str.replace('\n', '')
            if p:
                tmp_fn = fn.stem + '_unified.pickle'
                tmp_pn = dest_pick / tmp_fn
                df.to_pickle(tmp_pn)
                print(f'\nCreated {tmp_fn}\n')
            else:
                return res
        else:
            return dfs
    else:
        return df


def from_excel_ordinal(ordinal, _epoch0=datetime(1899, 12, 31)):
    # Convert Excel date shown as serial number into a date string
    if ordinal >= 60:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    return (_epoch0 + timedelta(days=ordinal)).replace(microsecond=0)


def df_2_xlsx_append(df, fn, sn, **kwargs):
    """
    =================================================

    🏷 Appends data from DataFrame to a new worksheet
      in existing file.

    ⚙ PREREQUISITES:
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    pip install pandas openpyxl

    📌 ARGUMENTS:
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    df  (DataFrame) your data
    fn  (Path)      Existing xlsx file
    sn  (str)       Sheet Name for a new data

    🎯 RETURNS
    ―――――――――――――――――――――――――――――――――――――――――――――――――
    → File list
    """

    #-----------------
    # Default values
    #-----------------
    ind = False

    # Get dynamic argument
    for k,v in kwargs.items():
        if k == 'index_on':
            ind = v

    # Alternative method - faster, shorter but no styling
    #------------------------------------------------------
    # with pd.ExcelWriter(fn, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
    # 
    #     if 'list' in str(type(df)):
    #         cr=0
    #         for d in df:
    #             d.to_excel(writer, sheet_name=sn, startrow=cr, index=ind)
    #             cr += d.shape[0]+2

    print(f'\nAppending data into file: {fn} | worksheet: {sn}\n')
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font,PatternFill
    import numpy as np

    wb = load_workbook(fn)
    ws = wb.create_sheet(sn)

    # Header row style
    fg_style = Font(color='00FFD966', bold=True)
    bg_style = PatternFill("solid", start_color="000d0d0d")

    if 'list' in str(type(df)):
        rc = 1 # row counter
        for d in df:
            for r in dataframe_to_rows(d, index=ind, header=True):
                ws.append(r)

                for y in range(1, d.shape[1]+1):
                    ws.cell(row=rc, column=y).font = fg_style
                    ws.cell(row=rc, column=y).fill = bg_style

            ws.append([np.nan])
            rc += d.shape[0]+2
    else:
        for r in dataframe_to_rows(df, index=ind, header=True):
            ws.append(r)

            for y in range(1, df.shape[1]+1):
                ws.cell(row=1, column=y).font = fg_style
                ws.cell(row=1, column=y).fill = bg_style
    try:
        wb.save(fn)
        print('Successfully append ✅\n')
    except:
        print('ERROR appending\n')

def df_2_xlsx(df, fn, sn, **kwargs):
    """
    ===================================================================

    🏷 Saves Dataframe(s) to worksheet(s) in a single Excel xlsx file.

    📌 ARGUMENTS:
    ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――
    - df             (DataFrame)  in multi mode [df1, df2, df3, ...]
    - fn             (Path(str))  File Name
    - sn             (str)        Sheet Name[s] in multi mode [sn1, sn2, sn3, ...]

    - tab_color    (hex color)  tab color
    - tab_colors   [hex colors] list of tab colors for every sheet
    - ac           (int)        0 = Off,    1 = On    (Auto-resize column)
    - style        (int)        0 = Header, 1 = Table
    - table_style  (str)        Name of Excel table style
    - properties   (dict)       custom file properties
    - index_on     (boolean)    Print with index True or False

    🎯 RETURNS:
    ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――
    → Single xlsx file with worksheet(s)

    """

    #-----------------
    # Default values
    #-----------------
    tc  = '#e6e6e6'
    if 'list' in str(type(df)):
        tcl = [tc]*len(df)
    ac  = 1
    s   = 0
    ind = False
    ts  = 'Table Style Medium 2'
    wsp = {
            'author':   'IgorP',
            'company':  'Private',
            'category': 'Report',
          }

    # Get dynamic argument
    for k,v in kwargs.items():
        if k == 'tab_color':
            tc = v
        if k == 'tab_colors':
            tcl = v
        if k == 'ac':
            ac = v
        if k == 'style':
            s = v
        if k == 'table_style':
            ts = v
        if k == 'properties':
            wsp = v
        if k == 'index_on':
            ind = v

    import xlsxwriter
    import pandas as pd

    def get_col_widths(dataframe):
        idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
        return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

    writer = pd.ExcelWriter(fn, engine='xlsxwriter')

    # Declare Excel Workbook
    workbook  = writer.book

    # Workbook properties
    workbook.set_properties(wsp)

    format_dict = {
        'bold'       : True,
        'text_wrap'  : True,
        'valign'     : 'middle',
        'font_color' : '#ffdf3c',
        'fg_color'   : '#232323',
        'border'     : 1
        }


    # Check if list of DataFrames is passed in argument
    if 'list' in str(type(df)):
        if len(df) == len(sn):
            for d,s, tagc in zip(df, sn, tcl):

                # Skip empty worksheets
                if d.empty:
                    pass
                else:
                    d.to_excel(writer, sheet_name=s, index=ind)

                    ws = writer.sheets[s]
                    for i, width in enumerate(get_col_widths(d)):
                        ws.set_column(i-1, i-1, width+2)

                    if s:
                        # Table style format
                        #--------------------
                        cols = []
                        for c in (d.columns.values):
                            cols.append({'header': c})

                        ws.add_table(0, 0, len(d.index), len(d.columns)-1, {'columns': cols, 'style': ts})

                    else:
                        # Header formatting (option instead of Table Style)
                        # -------------------------------------------------
                        header_format = workbook.add_format(format_dict)

                        for col_num, value in enumerate(d.columns.values):
                            ws.write(0, col_num, value, header_format)

                # Color the tabs
                ws.set_tab_color(tagc)
                # Freeze 1st row
                ws.freeze_panes(1, 0)
                # Worksheet zoom level
                ws.set_zoom(80)

        else:
            print('List of DataFrames and WorkSheets are not of the same size.')
            print('Exiting...')
            exit()
    # If not, check if the single DataFrame is actually a DataFrame type
    else:
        if 'DataFrame' in str(type(df)):

            # Create worksheet Data with data from dataframe
            df.to_excel(writer, sheet_name=sn, index=ind)

            # Cosmetic - auto-fit columns
            #-----------------------------------------------
            ws = writer.sheets[sn]

            if ac:
                for i, width in enumerate(get_col_widths(df)):
                    ws.set_column(i-1, i-1, width+2)

            # Style
            #------------------------------------
            if s:
                # Table style format
                #--------------------
                cols = []
                for c in (df.columns.values):
                    cols.append({'header': c})

                ws.add_table(0, 0, len(df.index), len(df.columns)-1, {'columns': cols, 'style': ts})

            else:
                # Header formatting (option instead of Table Style)
                # -------------------------------------------------
                header_format = workbook.add_format(format_dict)

                for col_num, value in enumerate(df.columns.values):
                    ws.write(0, col_num, value, header_format)

            # Color the tabs
            ws.set_tab_color(tc)
            # Freeze 1st row
            ws.freeze_panes(1, 0)
            # Worksheet zoom level
            ws.set_zoom(80)

        else:
            print('This is not a single dataframe for process.')
            print('Exiting...')
            exit()

    try:
        writer._save()
        print('\nSuccessfully saved: ',fn)
    except xlsxwriter.exceptions.FileCreateError:
        print('\n\nERROR!!!\nCannot write in opened file.\nCLOSE THE FILE, PLEASE!\n')


def print_df(df, **kwargs):
    """
    ====================================
    🏷 Prints a DataFrame.


    📌 ARGUMENTS:
    ――――――――――――――――――――――――――――――――――――
    🏁 FLAGS:  1 = ON, 0 = OFF

    df Dataframe         DataFrame
    d  DTypes            int
    dt DTypes tabular    int
    c  Columns           int
    v  Values            int
    vt Values tabular    int
            1 = simple
            2 = psql
    e  Exit after print  int

    🎯 RETURNS:
    ――――――――――――――――――――――――――――――――――――
    → DataFrame information and values
    """
    #-----------------
    # Default values
    #-----------------
    a_d  = 1
    a_dt = 0
    a_c  = 1
    a_v  = 1
    a_vt = 0
    a_e  = 1
    tf   = 0

    # Get dynamic argument
    for k,v in kwargs.items():

        if k == 'd':
            a_d  = v
        if k == 'dt':
            a_dt  = v
        if k == 'c':
            a_c  = v
        if k == 'v':
            a_v  = v
        if k == 'vt':
            a_vt  = v
        if k == 'e':
            a_e  = v

        if a_d==1 and a_dt==1:
            a_d  = 0

        if a_v==1 and a_vt>=1:
            a_v = 0


    print('\nDATAFRAME INFO | Rows:', df.shape[0], 'Columns:', df.shape[1])

    if a_c:
        print('Columns:',df.columns.tolist(),'\n')

    if a_d:
        print (df.dtypes,'\n')

    if a_dt:
        from tabulate import tabulate

        acc = []
        dtypes  = df.dtypes.tolist()
        columns = df.columns.tolist()

        for c,d in zip(columns, dtypes):
            acc.append([c,d])

        dfd = pd.DataFrame(acc, columns=['Column', 'Dtype'])

        print(tabulate(dfd,headers=dfd.columns,  tablefmt='psql',showindex=False),'\n')

    if a_vt:
        from tabulate import tabulate
        if a_vt == 1:
            tf='simple'
        elif a_vt == 2:
            tf='psql'
        elif a_vt == 2:
            tf='rounded_outline' # setx PYTHONIOENCODING="utf_8"
        print(tabulate(df, headers=df.columns.tolist(), tablefmt=tf ,showindex=False),'\n')

    if a_v:
        print('\n')
        for v in df.values.tolist():
            print(v)
    if a_e:
        exit()


def df_dtypes(df, mode):
    """
    =====================================

    🏷 Prints particular DataFrame types

    📌 ARGUMENTS:
    ――――――――――――――――――――――――――――――――――――
    - df    (DataFrame)
    - mode  (int)
        0 = print
        1 = NUM column list
        2 = DAT column list
        3 = TXT column list
        4 = ALL columns list

    🎯 RETURNS:
    ――――――――――――――――――――――――――――――――――――
    → Text data

    """

    tmp = []
    txt = []
    num = []
    dat = []

    dtypes  = df.dtypes.tolist()
    columns = df.columns.tolist()

    if mode == 0:
        for c,d in zip(columns, dtypes):
            print(d,'\t',c)
        exit()


    elif mode == 1:
        for c,d in zip(columns, dtypes):
            if d in ['float64', 'int64', 'int32']:
                print(d,'\t',c)
                num.append(c)
        return num

    elif mode == 2:
        for c,d in zip(columns, dtypes):
            if d in ['datetime64[ns]']:
                print(d,'\t',c)
                dat.append(c)
        return dat

    elif mode == 3:
        for c,d in zip(columns, dtypes):
            if d == 'O':
                print(d,'\t',c)
                txt.append(c)
        return txt

    else:
        for c,d in zip(columns, dtypes):
            tmp.append([c,d])
        return tmp


def split_df(df, lines = 1000):
    """
    ==================================================

    🏷 Splits DataFrame to list of smaller DataFrames

    📌 ARGUMENTS:
    ―――――――――――――――――――――――――――――
    - df (DataFrame)
    - lines    (int) - maximum lines per dataframe

    🎯 RETURNS
    ―――――――――――――――――――――――――――――
    → List of DataFrames
    """

    split_list = []
    num_chunks = len(df) // lines + 1

    for i in range(num_chunks):
        split_list.append(df[i*lines:(i+1)*lines])
    return split_list


def df_unique(df, col, id=0):
    res = df.loc[:,col].drop_duplicates()
    res.reset_index(drop=True, inplace=True)
    if id:
        res['ID'] = res.index
    return res


def clean_df(df, fillna=''):
    """
    ==================================

    🏷 Cleans DataFrame

    📌 ARGUMENTS:
    ――――――――――――――――――――――――――――――――――
    - df (DataFrame)
    - fillna   (str)

    🎯 RETURNS
    ――――――――――――――――――――――――――――――――――
    → Clean copy of original DataFrame
    """
    # Remove empty rows and columns
    res =  df.dropna(how='all', axis=0, inplace=False)
    res = res.dropna(how='all', axis=1, inplace=True)

    # Strip leading and trailing spaces in column names
    res.columns = res.columns.str.strip()
    # Ensure unique column names
    res.columns = [x[1] if x[1] not in res.columns[:x[0]] else f"{x[1]}_{list(res.columns[:x[0]]).count(x[1])}" for x in enumerate(res.columns)]

    # Strip leading and trailing spaces in all columns
    res = res.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    # Fillna
    if len(fillna):
        res.fillna(fillna,inplace=True)
    # Remove duplicates
    res = res.drop_duplicates()
    # Reindex
    res.reset_index(drop=True, inplace=True)

    return res
